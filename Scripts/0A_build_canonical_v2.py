# -*- coding: utf-8 -*-
"""
build_canonical_v2.py

Reads multi-year alfalfa phenotypic dataset from Excel (data_only mode to read evaluated formulas),
dynamically locates trait columns across inconsistent header structures, and exports a canonical dataset.
"""

from pathlib import Path
import openpyxl
import pandas as pd

# Define relative paths for data availability and reproducibility
BASE_DIR = Path(__file__).resolve().parent
RAW_DATA_PATH = BASE_DIR / "data" / "raw_phenotype_data.xlsx"
OUTPUT_DIR = BASE_DIR / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Dictionary for standardizing genotype / cultivar names
STANDARDIZE_GENOTYPE = {
    "Gongnog No.1": "Gongnong No.1",
    "Gongnong No.1": "Gongnong No.1",
    "Wlsouthern Spe": "WL Southern Special",
    "WlsouthernSpe": "WL Southern Special",
    "WlsouthernSpec": "WL Southern Special",
    "Algonguin": "Algonquin",
    "阿尔岗金": "Algonquin",
    "金皇后": "Golden Queen",
    "牧歌401+2": "AmeriGraze 401+Z",
    "巨人201+2": "AmeriStand 201+Z",
    "ACA542": "ACA542",
    "Xinjiang Daye": "Xinjiang_Daye",
}

def standardize_genotype(name: str) -> str:
    return STANDARDIZE_GENOTYPE.get(name, name)

def parse_float(val):
    try:
        return float(val)
    except (TypeError, ValueError):
        return float("nan")

# Load workbook in data_only mode to obtain evaluated formula values
wb = openpyxl.load_workbook(RAW_DATA_PATH, read_only=True, data_only=True)

records = []
sheet_years = [
    ("2002data ", 2002),
    ("2003data", 2003),
    ("2004data", 2004),
    ("2005data", 2005),
]

for sheet_name, year in sheet_years:
    if sheet_name not in wb.sheetnames:
        continue

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    # Locate header row index containing plot identifier ("小区号")
    hdr_idx = None
    for i, r in enumerate(rows[:6]):
        if any(v is not None and str(v).strip() == "小区号" for v in r):
            hdr_idx = i
            break

    if hdr_idx is None:
        continue

    def find_col(rows_to_search, keyword):
        for ri in rows_to_search:
            r = rows[ri]
            for ci, v in enumerate(r):
                if v is not None and keyword in str(v):
                    return ci
        return None

    # Dynamically locate trait column indices
    c_fw = find_col([hdr_idx + 1], "鲜重")

    c_dw = None
    r2 = rows[hdr_idx + 1]
    for ci, v in enumerate(r2):
        if v is not None and "风干重" in str(v) and "kg/kg" not in str(v) and "Kg/5m2" in str(v):
            c_dw = ci
            break

    c_ph = find_col([hdr_idx + 2], "平均")

    # Extract phenotype values
    for r in rows[hdr_idx + 3:]:
        if r[1] is None:
            continue
        plot_s = str(r[1]).strip()
        if not plot_s.isdigit():
            continue

        plot = int(plot_s)
        line = str(r[2]).strip() if r[2] is not None else ""
        rep = str(r[3]).strip() if len(r) > 3 and r[3] is not None else ""

        ph = parse_float(r[c_ph]) if c_ph is not None and c_ph < len(r) else float("nan")
        fw = parse_float(r[c_fw]) if c_fw is not None and c_fw < len(r) else float("nan")
        dw = parse_float(r[c_dw]) if c_dw is not None and c_dw < len(r) else float("nan")

        records.append((plot, line, rep, year, ph, fw, dw))

# Construct pandas DataFrame
df = pd.DataFrame(
    records,
    columns=[
        "plot",
        "genotype_original",
        "replicate",
        "calendar_year",
        "summer_height_cm",
        "fresh_weight_kg5m2",
        "dry_weight_kg5m2",
    ],
)

# Standardize IDs and compute experimental metadata
df["genotype_ID"] = df["genotype_original"].map(standardize_genotype)
df["cohort"] = df["plot"].apply(lambda p: "2001-established" if p <= 129 else "2002-established")
df["establishment_year"] = df["cohort"].apply(lambda c: 2001 if c == "2001-established" else 2002)
df["stand_age"] = df["calendar_year"] - df["establishment_year"]

# Export canonical wide-format dataset
output_file = OUTPUT_DIR / "canonical_cutting_wide_v2.csv"
df.to_csv(output_file, index=False, encoding="utf-8", na_rep="")

# Print summary
print(f"Total rows processed: {len(df)}")
print(f"Unique genotypes count: {df['genotype_ID'].nunique()}")
print(f"Yearly distribution: {df['calendar_year'].value_counts().sort_index().to_dict()}")
print(
    "Non-null trait counts - Summer Height: %d, Fresh Weight: %d, Dry Weight: %d"
    % (
        df["summer_height_cm"].notna().sum(),
        df["fresh_weight_kg5m2"].notna().sum(),
        df["dry_weight_kg5m2"].notna().sum(),
    )
)
